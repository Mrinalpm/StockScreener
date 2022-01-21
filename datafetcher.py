import csv
import os
import sys
from variables import *
from bselib.bse import BSE
import json
import tqdm
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from datetime import datetime
from dateutil.relativedelta import relativedelta

"""
Block 1 variables
"""
bse_codes = []

#----------------------------------------------------------------------------------------------------------
"""
Block 3 variables
"""
dict_containing_company_data = dict()

dict_containing_spreadsheet_headers_with_indexes = dict()

dict_containing_spreadsheet_headers = dict()

#The number of the right-most column. Will be updated when block 3 functions are called.
end_col = 0

list_with_headers_as_keys = []

#----------------------------------------------------------------------------------------------------------
"""
Block 1 functions
"""

"""
Purpose: To get the 6-digit BSE codes and store it in a global list

Prerequisites:
    * CSV file "active_equities_file_name" is up-to-date and existent in local directory.
    * "active_equities_file_name" contains BSE codes in column number column_number_containing_bse_codes

Updates:
    * Global variable "bse_codes" with all active equity codes.
"""
def get_bse_codes():
    try:
        csv_file = open(active_equities_file_name,'r')

        csv_file.readline()

        for a in csv.reader(csv_file, delimiter=','):
            bse_codes.append(a[column_number_containing_bse_codes-1])

    except FileNotFoundError as e:
        print("There is no file named " + active_equities_file_name + " at " + os.getcwd())
        sys.exit(1)
    except OSError:
        print("Could not open/read file: " +  os.getcwd() + active_equities_file_name)
        sys.exit(1)
    except Exception as err:
        print(f"Unexpected error opening {active_equities_file_name} is",repr(err))
        sys.exit(1)  # or replace this with "raise" ?

#----------------------------------------------------------------------------------------------------------
"""
Block 2 functions
"""

"""
Purpose: To get company data using bselib and store it in a file. THIS MAY TAKE A VERY LONG TIME

Prerequisites:
    * get_bse_codes()
    * All dictionary keys as described in variables.py matches what is returned by bselib

Creates:
    * "json_file_containing_data" containing company data
"""
def create_json_data():
    try:
        progressbar = tqdm.tqdm(bse_codes)
        b = BSE()
        data_dict = dict()
        count = 0

        for i in progressbar:
            local_dict = dict()

            if i not in data_dict:

                local_dict[QUOTE] = b.quote(i)

                local_dict[PERFORMANCE_ANALYSIS] = b.analysis(i)

                local_dict[FINANCIAL_RATIOS] = b.ratios(i)

                local_dict[PEER_COMPARISON] = b.peers(i)

                local_dict[CORPORATE_ACTIONS] = b.corporate_actions(i)

                local_dict[HOLDING_INFO_ANALYSIS] = b.holdings(i)

                data_dict[i] = local_dict #if all is well, write to main dict

                count += 1

                if count % company_data_write_frequency == 0:  #periodically write to file

                    file = open(json_file_containing_data, mode='w+')

                    all_of_it = file.read()

                    # close the file
                    file.close()

                    open(json_file_containing_data, 'w').close() #clear
                    try:
                        with open(json_file_containing_data, 'w') as fp:
                            json.dump(data_dict, fp, allow_nan=True)
                        fp.close()
                    except Exception: #something went wrong, restore old data
                        with open(json_file_containing_data, 'w') as fp:
                            fp.write(all_of_it)
                        fp.close()

                    count = 0
                progressbar.set_description("Reading companies")

        file = open(json_file_containing_data, mode='w+')

        all_of_it = file.read()

        # close the file
        file.close()

        open(json_file_containing_data, 'w').close() #clear
        try:
            with open(json_file_containing_data, 'w') as fp:
                json.dump(data_dict, fp, allow_nan=True)
            fp.close()
        except Exception: #something went wrong, restore old data
            with open(json_file_containing_data, 'w') as fp:
                fp.write(all_of_it)
            fp.close()
            
    except Exception as e:
        print(str(e))

#----------------------------------------------------------------------------------------------------------
"""
Block 3 functions
"""

"""
Purpose: To read company data from a .json file into the global dict

Prerequisites:
    * The file named "json_file_containing_data" must contain valid company data

Updates:
    * Global dictionary "dict_containing_company_data" eith full company data
"""
def read_from_json_file():
    global dict_containing_company_data
    try:
        with open(json_file_containing_data) as json_file:
            dict_containing_company_data = json.load(json_file)
    except Exception as e:
        print(str(e))

"""
Purpose: To insert a column within the dictionary "dict_containing_spreadsheet_headers_with_indexes"
        An extra column is inserted AFTER the specified "column" parameter

Updates:
    * The dictionary "dict_containing_spreadsheet_headers_with_indexes" with an added column after "column"
"""
def insert_column_within_dict(column):
    global dict_containing_spreadsheet_headers_with_indexes
    for row in dict_containing_spreadsheet_headers_with_indexes:
        for col in dict_containing_spreadsheet_headers_with_indexes[row]:
            insert_column_within_dict_helper(col, dict_containing_spreadsheet_headers_with_indexes)

"""
Purpose: To insert a column within the dictionary "dict_containing_spreadsheet_headers_with_indexes"
        An extra column is inserted AFTER the specified "column" parameter
        Note: THIS IS A HELPER METHOD. insert_column_within_dict() IS THE ONLY METHOD THAT MUST CALL THIS

Updates:
    * The dictionary "dict_containing_spreadsheet_headers_with_indexes" with an added column after "column"
"""
def insert_column_within_dict_helper(col, my_dict):
    next_col = str(int(col) + 1)
    if next_col in my_dict:
        insert_column_within_dict_helper(next_col, my_dict)
        my_dict[next_col] = my_dict[col]
    else:
        my_dict[next_col] = my_dict[col]

"""
Purpose: To insert a header value within the dictionary "dict_containing_spreadsheet_headers_with_indexes"

Updates:
    * The dictionary "dict_containing_spreadsheet_headers_with_indexes" with the new value
"""
def insert_dict_containing_spreadsheet_headers(row, column, value):
    global dict_containing_spreadsheet_headers_with_indexes, end_col
    if row in dict_containing_spreadsheet_headers_with_indexes:
        if not column in dict_containing_spreadsheet_headers_with_indexes[row]:
            dict_containing_spreadsheet_headers_with_indexes[row][column] = value
            if column > end_col:
                end_col = column
    else:
        dict_containing_spreadsheet_headers_with_indexes[row] = dict()
        dict_containing_spreadsheet_headers_with_indexes[row][column] = value
        if column > end_col:
            end_col = column

"""
Purpose: Populates the headers in dictionary "dict_containing_spreadsheet_headers_with_indexes"

Prerequisites:
    * "dict_containing_company_data" must be properly initialized with company data

Updates:
    * "dict_containing_spreadsheet_headers_with_indexes" with updated header values and corresponding indexes.
"""
def create_header_position():
    global dict_containing_spreadsheet_headers, dict_containing_spreadsheet_headers_with_indexes
    temp_dict = {}
    convert(dict_containing_spreadsheet_headers, temp_dict)
    create_header_position_helper(temp_dict, 0, 0)

"""
Purpose: Populates the headers in dictionary "dict_containing_spreadsheet_headers_with_indexes"
        Note: THIS IS A HELPER METHOD. create_header_position() IS THE ONLY METHOD THAT MUST CALL THIS

Prerequisites:
    * "dict_containing_company_data" must be properly initialized with company data

Updates:
    * "dict_containing_spreadsheet_headers_with_indexes" with updated header values and corresponding indexes.
"""
def create_header_position_helper(company_data, row, column):
    """"""
    offset = 0
    if isinstance(company_data, dict):
        for key in company_data:
            insert_dict_containing_spreadsheet_headers(row, column + offset, key)
            offset += create_header_position_helper(company_data[key], row + 1, column + offset)

            if not isinstance(company_data[key], dict):
                offset += 1

            if isinstance(company_data[key], dict) and not company_data[key]:
                offset += 1
    return offset

"""
Purpose: To convert a dictionary/list into a nesteddictionary. This includes a dictionary with nested lists/dictionaries.

Parameters:
    * company_data: The dictionary containing nested lists/dictionary
    *to_insert: The new dictionary that will be created. This will not have any lists

Updates:
    *to_insert: With no lists, only a purely nested dictionary
"""
def convert(company_data, to_insert):
    if isinstance(company_data, dict):
        for key in company_data:
            if isinstance(company_data[key], list) or isinstance(company_data[key], dict):
                to_insert[key] = {}
                convert(company_data[key], to_insert[key])
            else:
                to_insert[key] = company_data[key]
    elif isinstance(company_data, list):
        for entry in range(0, len(company_data)):
            if isinstance(company_data[entry], list) or isinstance(company_data[entry], dict):
                to_insert[entry] = {}
                convert(company_data[entry], to_insert[entry])
            else:
                to_insert[entry] = company_data[entry]

"""
Purpose: To create a mega dictionary containing data of ALL companies. Older data will get overwritten as newer columns get detected.
        This method is mainly used for gathering the headers for the spreadsheet

Prerequisites:
    * "dict_containing_company_data" must be properly initialized with company data

Updates:
    * "dict_containing_spreadsheet_headers" with updated header values
"""
def create_mega_data_dict():
    global dict_containing_company_data, dict_containing_spreadsheet_headers

    progressbar = tqdm.tqdm(dict_containing_company_data)
    progressbar.set_description("Gathering data ...")
    for company_code in progressbar:
        company_data_dict = dict_containing_company_data[company_code]
        merge(company_data_dict, dict_containing_spreadsheet_headers)

"""
Purpose: To merge two dictionaries. This also takes care of any nested dictionaries and nested lists that may exist

Updates:
    * header_data: The second dictionary that gets passed gets updated with the combined entries of both dictionaries.
"""
def merge(company_data, header_data):
    if isinstance(company_data, dict):
        for key in company_data:
            if not key in header_data:
                header_data[key] = company_data[key]
            else:
                if isinstance(company_data[key], dict) or isinstance(company_data[key], list) or isinstance(header_data[key], dict) or isinstance(header_data[key], list):
                    merge(company_data[key], header_data[key])
    elif isinstance(company_data, list):
        if len(header_data) < len( company_data):
            og_len = len(header_data)
            for index in range(og_len, len(company_data)):
                header_data.append(company_data[index])

"""
Purpose: To get the size of a data structure. This includes nested dictionaries, nested lists and any combinations of the two.

Return: The size of the data structure
"""
def get_data_size(my_data):
    return get_data_size_helper(my_data, 0)

"""
Purpose: To get the size of a data structure. This includes nested dictionaries, nested lists and any combinations of the two.
        Note: THIS IS A HELPER METHOD. get_data_size(my_data) IS THE ONLY METHOD THAT MUST CALL THIS
Return: The size of the data structure
"""
def get_data_size_helper(my_data, size):
    current_size = size
    if isinstance(my_data, dict):
        for key in my_data:
            if isinstance(my_data[key], dict) or isinstance(my_data[key], list):
                current_size += get_data_size_helper(my_data[key], size)
            else:
                current_size += 1
    elif isinstance(my_data, list):
        current_size += len(my_data)
    
    return current_size

"""
Purpose: To write the headers to the appropriate spreadsheet

Prerequisites:
    * "dict_containing_spreadsheet_headers_with_indexes" must be properly initialized with the proper headers and corresponding indexes.
"""
def write_headers_to_spreadsheet():
    global dict_containing_spreadsheet_headers_with_indexes
    wb = openpyxl.Workbook()

    print("Preparing spreadsheet headers ...")

    # grab the active worksheet
    ws = wb.active

    for row_i in dict_containing_spreadsheet_headers_with_indexes:
        for col_i in dict_containing_spreadsheet_headers_with_indexes[row_i]:
            ws.cell (row = row_i + 1, column = col_i + 1).value = str(dict_containing_spreadsheet_headers_with_indexes[row_i][col_i])

    wb.save(SPREADSHEET_FILE_NAME)
    wb.close()

"""
Purpose: To fill in any missing headers AFTER write_headers_to_spreadsheet() has been called

Prerequisites:
    * write_headers_to_spreadsheet() must be called
    * "dict_containing_spreadsheet_headers_with_indexes" must be properly initialized with the proper headers and corresponding indexes.
"""
def write_missing_headers_to_spreadsheet():
    global end_col, list_with_headers_as_keys

    wb_obj = openpyxl.load_workbook(SPREADSHEET_FILE_NAME)
    sheet_obj = wb_obj.active

    for col in range (1, end_col + 1):
        start_row = 1
        for row_i in range (len(dict_containing_spreadsheet_headers_with_indexes), 0, -1):
            cell_obj = sheet_obj.cell(row = row_i, column = col).value
            if not cell_obj is None:
                start_row = row_i
                break

        for row_i in range (start_row, 0, -1):
            if sheet_obj.cell(row = row_i, column = col).value is None:
                col_value = col - 1
                new_value = sheet_obj.cell(row = row_i, column = col_value).value
                while new_value is None:
                    new_value = sheet_obj.cell(row = row_i, column = col_value).value
                    col_value -= 1
                sheet_obj.cell (row = row_i, column = col).value = str(new_value)

    wb_obj.save(SPREADSHEET_FILE_NAME)

    for col in range (1, end_col + 1):
        my_list = []
        for row_i in range (1, len(dict_containing_spreadsheet_headers_with_indexes) + 1):
            value = sheet_obj.cell(row = row_i, column = col).value
            if not value is None:
                my_list.append(value)
            else:
                break
        list_with_headers_as_keys.append(my_list)

    wb_obj.close()

"""
Purpose: To write company data to the spreadsheet

Prerequisites:
    * write_missing_headers_to_spreadsheet() must be called first
    * dict_containing_company_data must be initialized with company data
"""
def write_company_data_to_spreadsheet():
    global list_with_headers_as_keys, dict_containing_company_data, dict_containing_spreadsheet_headers_with_indexes

    wb_obj = openpyxl.load_workbook(SPREADSHEET_FILE_NAME)
    sheet_obj = wb_obj.active

    progressbar = tqdm.tqdm(dict_containing_company_data)
    progressbar.set_description("Writing company data ...")
    row = len(dict_containing_spreadsheet_headers_with_indexes) + 2

    sheet_obj.cell (row = 1, column = end_col + 1).value = "Dividend yield"
    sheet_obj.cell (row = 1, column = end_col + 2).value = "5 yr. avg. dividend yield"

    for company_code in progressbar:
        company_data = dict_containing_company_data[company_code]
        dividend_info = get_dividend_info(company_data)
        for col in range (1, end_col + 1):
            data = get_data_value(company_data, list_with_headers_as_keys[col - 1])
            if data is None:
                sheet_obj.cell (row = row, column = col).value = str("")
            else:
                sheet_obj.cell (row = row, column = col).value = ILLEGAL_CHARACTERS_RE.sub(r'',str(data))
        if dividend_info[0] is None:
            sheet_obj.cell (row = row, column = end_col + 1).value = "0.0"
        else:
            sheet_obj.cell (row = row, column = end_col + 1).value = str(dividend_info[0])

        if dividend_info[1] is None:
            sheet_obj.cell (row = row, column = end_col + 2).value = "0.0"
        else:
            sheet_obj.cell (row = row, column = end_col + 2).value = str(dividend_info[1])

        row += 1

    print("Saving file ...")
    wb_obj.save(SPREADSHEET_FILE_NAME)
    wb_obj.close()

"""
Returns a look up table in the form of a dictionary
Input: A list containing the headers
Return: A dictionary containing corresponding indexes for each of the header values
"""
def get_lookup_table(param_list):
    to_return = dict()

    for i in range(len(param_list)):
        to_return[param_list[i]] = i

    return to_return

def get_dividend_info(company_data):

    #Values to be returned
    fiveyryield = None
    currentyield = None
    stock_price = None
    face_value = None

    if QUOTE_STOCK_PRICE in company_data[QUOTE]:
        stock_price = validate_double(company_data[QUOTE][QUOTE_STOCK_PRICE])

    if QUOTE_FACE_VALUE in company_data[QUOTE]:
        face_value = validate_double(company_data[QUOTE][QUOTE_FACE_VALUE])

    if stock_price and face_value:
        div_data = company_data[CORPORATE_ACTIONS]
        end_of_fin_yr = datetime.now()

        # each of the previous 5 year slabs. This will contain dividend percentages
        prev_yr_totals = [0.0, 0.0, 0.0, 0.0, 0.0]

        if CORPORATE_ACTIONS_DIVIDENDS in div_data:
            if CORPORATE_ACTIONS_DIVIDENDS_DATA in div_data[CORPORATE_ACTIONS_DIVIDENDS]:
                dictionary = get_lookup_table(div_data[CORPORATE_ACTIONS_DIVIDENDS][CORPORATE_ACTIONS_DIVIDENDS_HEADER])  # get the lookup table
                for i in div_data[CORPORATE_ACTIONS_DIVIDENDS][CORPORATE_ACTIONS_DIVIDENDS_DATA]:
                    date = datetime.strptime(i[dictionary[CORPORATE_ACTIONS_DIVIDENDS_RECORD_DATE]], date_format)  # date of dividend issued

                    if end_of_fin_yr - relativedelta(years=1) < date <= end_of_fin_yr:
                        prev_yr_totals[0] += float(i[dictionary[CORPORATE_ACTIONS_DIVIDENDS_DIVIDEND_PERCENTAGE]].replace('%', ''))
                        # dividend issued current year
                    elif end_of_fin_yr - relativedelta(years=2) < date <= end_of_fin_yr - relativedelta(years=1):
                        prev_yr_totals[1] += float(i[dictionary[CORPORATE_ACTIONS_DIVIDENDS_DIVIDEND_PERCENTAGE]].replace('%', ''))
                        # dividend issued previous year
                    elif end_of_fin_yr - relativedelta(years=3) < date <= end_of_fin_yr - relativedelta(years=2):
                        prev_yr_totals[2] += float(i[dictionary[CORPORATE_ACTIONS_DIVIDENDS_DIVIDEND_PERCENTAGE]].replace('%', ''))
                        # dividend issued two years ago
                    elif end_of_fin_yr - relativedelta(years=4) < date <= end_of_fin_yr - relativedelta(years=3):
                        prev_yr_totals[3] += float(i[dictionary[CORPORATE_ACTIONS_DIVIDENDS_DIVIDEND_PERCENTAGE]].replace('%', ''))
                        # dividend issued three years ago
                    elif end_of_fin_yr - relativedelta(years=5) < date <= end_of_fin_yr - relativedelta(years=4):
                        prev_yr_totals[4] += float(i[dictionary[CORPORATE_ACTIONS_DIVIDENDS_DIVIDEND_PERCENTAGE]].replace('%', ''))
                        # dividend issued four years ago

                # five year yield returned as a percentage
                fiveyryield = sum(prev_yr_totals) * float(face_value) / (5.0 * float(stock_price))
                currentyield = prev_yr_totals[0] * float(face_value) / float(stock_price)  # current yield returned as a percentage

    return currentyield, fiveyryield

"""
Function for validating a float as a string
Returns None if invalid
"""
def validate_double(str_param):

    str_param = str(str_param)
    # noinspection PyBroadException
    try:
        to_return = str(float(str_param.replace(',', '')))  # if any commas exist, remove them
    except Exception:  # exception occured, just return NULL
        to_return = None
    return to_return
"""
Purpose: To extract a value from a nested dictionary/list based on a key list

Parameters: 
    * company_data: Contains data
    * header_list: The list containing keys/headers
"""
def get_data_value(company_data, header_list):
    if (not company_data is None) and len(header_list) > 0 and len (company_data) > 0:
        if isinstance (company_data, dict):
            if header_list[0] in company_data:
                return get_data_value(company_data[header_list[0]], header_list[1:])
            else:
                return None
        elif isinstance (company_data, list):
            array_index = int(header_list[0])
            if array_index < len(company_data):
                return get_data_value(company_data[array_index], header_list[1:])
            else:
                return None
        else:
            if (isinstance(company_data, list) or isinstance(company_data, dict)) and not company_data:
                return None
            else:
                return company_data
    else:
        if (isinstance(company_data, list) or isinstance(company_data, dict)) and not company_data:
                return None
        else:
            return company_data
 
#----------------------------------------------------------------------------------------------------------

"""
Only execute block 1 and 2 functions if latest data is desired. this can take up to 24 hours.
If you only want the spreadsheet generated based on an existing json, run only block 3 functions
"""

"""
Block 1
"""
#get_bse_codes()

"""
Block 2
"""
#create_json_data()

"""
Block 3
"""
read_from_json_file()
create_mega_data_dict()
create_header_position()
write_headers_to_spreadsheet()
write_missing_headers_to_spreadsheet()
write_company_data_to_spreadsheet()